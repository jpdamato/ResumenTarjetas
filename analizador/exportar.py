"""Exporta a Excel (.xlsx) el detalle de movimientos que se ve en la tabla.

Aplica los MISMOS filtros que el tablero (moneda, rango de periodos, banco,
titular, categoria, busqueda) y, como todo lo demas, va acotado al usuario: se
exporta solo lo suyo. A diferencia de la tabla en pantalla —que corta en 600
filas por prolijidad— la exportacion trae TODAS las filas del filtro.

Los pagos se excluyen, igual que en el analisis: no son gasto.
"""

from __future__ import annotations

import datetime as dt
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MONEDAS = ("ARS", "USD")

# (titulo, ancho, alineacion). El orden es el de las columnas del Excel.
COLUMNAS = [
    ("Fecha", 12, "left"),
    ("Período", 10, "left"),
    ("Descripción", 40, "left"),
    ("Comercio", 26, "left"),
    ("Categoría", 20, "left"),
    ("Cuota", 8, "center"),
    ("Titular", 22, "left"),
    ("Banco", 16, "left"),
    ("Moneda", 8, "center"),
    ("Tipo", 12, "left"),
    ("Importe", 16, "right"),
]

_TIPO = {"purchase": "Compra", "cost": "Costo tarjeta", "payment": "Pago"}


def _like(texto: str) -> str:
    """Escapa los comodines de LIKE para que la busqueda sea substring literal.

    Sin esto, buscar "100%" o "a_b" en el tablero traeria de mas: '%' y '_' son
    comodines de SQL. Se escapan con '\\' (declarado con ESCAPE en el LIKE).
    """
    escapado = texto.lower().replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
    return f"%{escapado}%"


def filtrar_movimientos(conn, usuario_id: int, f: dict) -> list:
    """Filas del usuario que cumplen el filtro, en orden de resumen.

    Reproduce en SQL lo que hace filtrar() en el tablero (template.html): mismo
    criterio, para que el Excel contenga exactamente lo que se esta viendo.
    """
    moneda = f.get("moneda") if f.get("moneda") in MONEDAS else "ARS"
    cond = ["usuario_id = ?", "tipo != 'payment'", "moneda = ?"]
    params: list = [usuario_id, moneda]

    for campo, columna in (("desde", "periodo >= ?"), ("hasta", "periodo <= ?"),
                           ("banco", "banco = ?"), ("titular", "titular = ?"),
                           ("categoria", "categoria = ?")):
        if f.get(campo):
            cond.append(columna)
            params.append(f[campo])

    if f.get("q"):
        cond.append("(LOWER(descripcion) LIKE ? ESCAPE '\\' "
                    "OR LOWER(comercio) LIKE ? ESCAPE '\\')")
        params += [_like(f["q"]), _like(f["q"])]

    sql = ("SELECT * FROM movimientos WHERE " + " AND ".join(cond)
           + " ORDER BY periodo, orden")
    return conn.execute(sql, params).fetchall()


def _fecha(valor) -> dt.date | str:
    """La fecha se guarda como texto ISO; se pasa a date real para que Excel la
    trate como fecha (y se ordene/filtre como tal), no como texto."""
    try:
        return dt.date.fromisoformat(valor) if valor else ""
    except (ValueError, TypeError):
        return valor or ""


def generar_xlsx(conn, usuario_id: int, f: dict, usuario: str) -> bytes:
    filas = filtrar_movimientos(conn, usuario_id, f)
    moneda = f.get("moneda") if f.get("moneda") in MONEDAS else "ARS"

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    negrita = Font(bold=True)
    fondo = PatternFill("solid", fgColor="EEECE4")
    fmt_importe = f'#,##0.00;[Red]-#,##0.00'

    for i, (titulo, ancho, _al) in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=1, column=i, value=titulo)
        c.font = negrita
        c.fill = fondo
        ws.column_dimensions[get_column_letter(i)].width = ancho

    total = 0.0
    fila = 2
    for m in filas:
        cuota = f"{m['cuota_nro']}/{m['cuota_total']}" if m["cuota_nro"] else ""
        valores = [
            _fecha(m["fecha"]), m["periodo"], m["descripcion"], m["comercio"],
            m["categoria"], cuota, m["titular"] or "", m["banco"],
            m["moneda"], _TIPO.get(m["tipo"], m["tipo"]), round(m["importe"], 2),
        ]
        for col, (valor, (_t, _w, al)) in enumerate(zip(valores, COLUMNAS), start=1):
            c = ws.cell(row=fila, column=col, value=valor)
            c.alignment = Alignment(horizontal=al)
            if col == len(COLUMNAS):
                c.number_format = fmt_importe
            elif col == 1 and isinstance(valor, dt.date):
                c.number_format = "dd/mm/yyyy"
        total += float(m["importe"])
        fila += 1

    # Fila de total, alineada con el pie de la tabla del tablero.
    etiqueta = ws.cell(row=fila, column=len(COLUMNAS) - 1,
                       value=f"Total ({len(filas)} movimientos)")
    etiqueta.font = negrita
    etiqueta.alignment = Alignment(horizontal="right")
    tot = ws.cell(row=fila, column=len(COLUMNAS), value=round(total, 2))
    tot.font = negrita
    tot.number_format = fmt_importe

    ws.freeze_panes = "A2"           # el encabezado queda fijo al hacer scroll
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}1"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def nombre_archivo(usuario: str, f: dict) -> str:
    """p. ej. movimientos-jpd-2025-01_2026-07-ARS.xlsx"""
    limpio = "".join(c for c in usuario if c.isalnum() or c in "._-") or "tarjetas"
    partes = ["movimientos", limpio]
    if f.get("desde") or f.get("hasta"):
        partes.append(f"{f.get('desde', '')}_{f.get('hasta', '')}")
    partes.append(f.get("moneda") if f.get("moneda") in MONEDAS else "ARS")
    return "-".join(partes) + ".xlsx"
